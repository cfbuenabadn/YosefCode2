from brainSources import cortical_sourceFolders, olfactory_sourceFolders, bateup_sourceFolders, samIsrael_sourceFolders
import os
import os.path
import glob
import openpyxl
import re
from collections import namedtuple

#The sourceFolders is a tuple of dir and bool (isPairedEnd) - get rid of the latter field which is not needed here

SOURCES_DIR = "/data/yosef/BRAIN/sources"
#PROCESSED_DIR = "/data/yosef/BRAIN/processed_June2015_b"
#PROCESSED_DIR = "/data/yosef/BRAIN/processed_July2015"
#PROCESSED_DIR = "/data/yosef/BRAIN/processed_Bateup_Aug2015"
#PROCESSED_DIR = "/data/yosef/BRAIN/processed_Sep2015"
PROCESSED_DIR = "/data/yosef/BRAIN/processed_Zebrafish_Oct2015/"

#Cortical / Olfactory / Bateup / SamIsrael
PROJECT = "SamIsrael"
if PROJECT == "Cortical":
    sourceFolders = cortical_sourceFolders
    METADATA_DIR = "/data/yosef/BRAIN/sources/metadata/cortical"
    CONFIG_OUTPUT_FILE = os.path.join(PROCESSED_DIR, 'collect/config_cortical.xlsx')
elif PROJECT == "Olfactory":
    sourceFolders = olfactory_sourceFolders
    METADATA_DIR = "/data/yosef/BRAIN/sources/metadata/olfactory"
    CONFIG_OUTPUT_FILE = os.path.join(PROCESSED_DIR, 'collect/config_olfactory.xlsx')
elif PROJECT == "Bateup":
    sourceFolders = bateup_sourceFolders
    METADATA_DIR = "/data/yosef/BRAIN/sources/metadata/bateup"
    CONFIG_OUTPUT_FILE = os.path.join(PROCESSED_DIR, 'collect/config_bateup.xlsx')
elif PROJECT == "SamIsrael":
    sourceFolders = samIsrael_sourceFolders
    METADATA_DIR = "/data/yosef/BRAIN/sources/metadata/samisrael"
    CONFIG_OUTPUT_FILE = os.path.join(PROCESSED_DIR, 'collect/config_samisrael.xlsx')
else:
    raise Exception("unrecognized project!")



sourceFolders = [x[0] for x in sourceFolders]

#observe that the cuff cell list and the rsem cell list (and in the future kallisto cell list?!) should all be the same
outputCellListFileName = os.path.join(PROCESSED_DIR, 'collect/cuff/cell_list.txt')

MD_FIELDS = [fieldname.replace(' ', '_').lower() for fieldname in \
    ['Expt_condition', 'sample_id', 'sample_sequencing_id', 'sequencing_run_id', 'C1_run_id', 'Single_Bulk_Pool', 'Dilution plate coordinate', 'C1 position', 'library plate name', 'library plate position', 'Barcode', 'Index set', 'Lane', 'cDNA concentration', 'number_of_animals', 'DOB', 'Sex', 'Age', 'FACS_C1_date', 'C1_chip', 'Cell type', 'Cell description', 'C1', 'Capture program temp', 'spikein_dilution']]
md_record = namedtuple('md_record', MD_FIELDS)



def ReadOneMetadataFile(md_file):
    wb = openpyxl.load_workbook(filename=md_file, data_only=True, read_only=True)
    ws = wb.active #get active worksheet
    ws_rows_iterator = ws.rows

    #ws.rows is an iterator over the worksheet's rows
    #the if cell makes sure that we're not reading empty cells
    header_row = [cell.value.encode("ascii").strip().replace(' ', '_').lower() if cell.value else "NO_HEADER_VALUE_IN_CELL" for cell in ws_rows_iterator.next()]

    #Sam Israel adds more metadata after the standard colums. In the meantime - solve this by making away with these columns
    header_row = header_row[0:len(MD_FIELDS)]

    #make sure that the excel contains the expected columns
    if(cmp(header_row, MD_FIELDS) != 0):
        for col in xrange(0, len(MD_FIELDS)):
            if header_row[col] != MD_FIELDS[col]:
		        raise Exception("metadata excel file (%s) is not in the expected format\nColumn %d differs %s vs. %s" % md_file, col, header_row[col], MD_FIELDS[col])
	raise Exception("metadata excel file (%s) is not in the expected format\nNot the same number of columns" % md_file)


    md_dict_for_file = {}

    #read the rest of the rows:
    count = 0
    for row in ws_rows_iterator:
        #I convert cell.value to unicode string because openpyxl reads int vals as ints, dates as dates etc.
        #(at first I converted to str(cell.value), which converts to an ascii string, but then I found that the degree symbol Dave uses in the  temperature field is unicode, so I cannot convert it to ascii)
        row_metadata = [unicode(cell.value).strip().replace(' ', '_').upper() for cell in row]

        count += 1
        #print "read row %d" % count #for debug

        #empty cells will become the string "NONE" replace those with "NA" for consistency:
        row_metadata = ["NA" if x == "NONE" else x for x in row_metadata]

        #Sam Israel adds more metadata after the standard colums. In the meantime - solve this by making away with these columns
        row_metadata = row_metadata[0:len(MD_FIELDS)]


        cur_cell_md_record = md_record._make(row_metadata)

        if(all(r == "NA" for r in row_metadata)):
            #this happens for some reason in some of the excel files - an entire row of Nones (that I transform to "NA" earlier in the loop) => maybe they correspond to an extra newline, i.e., empty line, in the excels?
            continue

        if not(cur_cell_md_record.sample_id) or not(cur_cell_md_record.sample_sequencing_id):
            raise Exception("cannot have an empty cell id")

        md_dict_for_file[cur_cell_md_record.sample_sequencing_id] = cur_cell_md_record




    return md_dict_for_file

def ReadMetadata():
    md_files = glob.glob(os.path.join(METADATA_DIR, '*.xlsx')) + glob.glob(os.path.join(METADATA_DIR, '*.xls'))
    md_dict = {}

    for md_file in md_files:
        print "reading metadafile " + md_file
        partial_md_dict = ReadOneMetadataFile(md_file)

        if md_dict and frozenset(partial_md_dict.keys()).intersection(md_dict.keys()):
            # if md_dict then the dictionary is not empty (this is not the first iteration
            raise Exception("Reading metadata file: The metadata dictionaries' keys should be mutually exclusive - they are indexed by the unique_id of the cell")

        #merge the partial dictionary into the aggregate dictionary
        md_dict.update(partial_md_dict)

    return md_dict


def ProcessOutputCellList(outputCellListFileName):
    cellID2outputName_dict = {}

    outputCellList = [row.strip() for row in open(outputCellListFileName).readlines()]

    #make the output cell list into a dictionary where the cells unique id points to its dir structure
    #the unique ID is non-greedy to make sure it does not swallow the optional barcode
    #Shana unique IDs were added after her change to the output format on Oct 2015
    extractUniqueIDFromCellName = re.compile(r"^(?P<path>[\d\w_\-]+)/(?P<uniqueID>[\w_]+?)(_[ACGT]+\-[ACGT]+)?(_(?P<shana_unique_id>S(\d{1,4})))?(_(?P<lane>L\d\d\d))?$")
    for cell in outputCellList:

        m = re.match(extractUniqueIDFromCellName, cell)
        #print(cell) #for debug
        if not(m.groupdict().has_key("path")) or not(m.groupdict().has_key("uniqueID")):
            raise Exception("Reading collect output list: Cannot parse cell name from the output list! (cell name: %s)" % cell)

        cell_uniqueID = m.group("uniqueID")
        #quick and dirty solution for the batch with the bad naming format
        #--> no longer needed after Shana revised the output format according to my requests, Oct 2015
        if(False and m.group("path").startswith("150904_JGI1")):
            cell_uniqueID += "_" + m.group("lane")

        if(cellID2outputName_dict.has_key(cell_uniqueID)):
            raise Exception("Reading collect output list: Same cell encountered twice in the output?! (cell name: %s)" % cell)

        cellID2outputName_dict[cell_uniqueID] = cell


    return cellID2outputName_dict


def WriteConfigFile(elementsToWrite):
    #the values in the element list are: (cellName, metadata_for_cell, outputName_for_cell)

    #use only the values, sort them by the cell's unique id which is the first element in the tuple
    elementsToWrite = sorted(elementsToWrite.values(), key=lambda x: x[0])

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()

    #compose and print the header line - it is made of the header fields, I just want to place the unique ID first before all the rest of the fields
    #and prefix every MD on the cell with "MD_"
    headerLineColumns = [MD_FIELDS[1], MD_FIELDS[2], 'output_name'] + ["MD_" + MD_FIELDS[0]] + ["MD_" + x for x in MD_FIELDS[3:]]
    ws.append(headerLineColumns)

    for (cellName, metadata_for_cell, outputName_for_cell) in elementsToWrite:
        cellColumns = [metadata_for_cell[1], metadata_for_cell[2], outputName_for_cell]
        for field_ind in [0] + range(3, len(MD_FIELDS)):
            cellColumns.append(metadata_for_cell[field_ind])

        ws.append(cellColumns)

    wb.save(CONFIG_OUTPUT_FILE)

    #consider writing also as a tab delimited file because R's read.xlsx is soooo slow
    #---> in the meantime I use in R read.xlsx2




output_cell_dict = ProcessOutputCellList(outputCellListFileName)

metadata_dict = ReadMetadata()

elementsToWrite = {}
allCellsInAllFolders = [] #elementsToWrite.keys() is not equivalent to the list of all cells for which there is a source folder because I do not add a uniqueID to that list if there's no metadata file for it, for example
for sourceFolder in [os.path.join(SOURCES_DIR, folder) for folder in sourceFolders]:
    print "reading sources folder: %s" % sourceFolder
    cellsInFolder = sorted([d for d in os.listdir(sourceFolder) \
                     if os.path.isdir(os.path.join(sourceFolder, d)) and d.lower() != "Undetermined_indices".lower() and d.lower() != "home".lower()])

    #remove the prefix "Sample_" from the cell name if it is there
    cellsInFolder = [cellName[len("Sample_"):] if cellName.startswith("Sample_") else cellName for cellName in cellsInFolder]

    allCellsInAllFolders += cellsInFolder

    for cellName in cellsInFolder:

        metadata_for_cell = metadata_dict.get(cellName)
        if not(metadata_for_cell):
            print "Reading source data directories: Warning: cell %s has no metadata associated with it - skipping it..." % cellName
            continue
        else:
        #   print "Read metadata for cell %s" % cellName
            pass

        outputName_for_cell = output_cell_dict.get(cellName)
        if not(outputName_for_cell):
            print "Reading source data directories: Warning: cell %s has no collected output associated with it - skipping it..." % cellName
            continue

        if(elementsToWrite.has_key(cellName)):
            raise Exception("Reading source data directories: unique cell ID (%s) encountered twice?!")

        #the value associated with the key is a tuple: the uniqueID, its metadata and its output name
        elementsToWrite[cellName] = (cellName, metadata_for_cell, outputName_for_cell)

    print "done reading sources folder: %s" % sourceFolder
    print "*********************************************\n\n"


numCellDirectoriesRead = len(allCellsInAllFolders)
allCellsInAllFolders = frozenset(allCellsInAllFolders)
if(numCellDirectoriesRead > len(allCellsInAllFolders)):
    raise Exception("The list of unique IDs read from the source folders is not unique...")

#now, see if there are any metadata entries for which no file was found
unclaimed_metadata_entries = sorted(frozenset(metadata_dict.keys()).difference(allCellsInAllFolders))
for entry in unclaimed_metadata_entries:
    print "Warning: cell %s appeared in metadata list but had no source directory associated with it - skipping it..." % entry

#now, see if there are any collect output entries for which no file was found
unclaimed_output_entries = sorted(frozenset(output_cell_dict.keys()).difference(allCellsInAllFolders))
for entry in unclaimed_output_entries:
    print "Warning: cell %s appeared in collect output list but had no source directory associated with it - skipping it..." % entry

#you collected all the elements to be written, now write the config file
WriteConfigFile(elementsToWrite)

print "all done"